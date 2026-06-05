import os
import sys
import uuid
import base64
from datetime import datetime
from typing import List, Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openpyxl
import traceback
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import queue
import threading

HAS_TKINTER = True
try:
    import tkinter as tk
    from tkinter import filedialog
except (ImportError, ModuleNotFoundError):
    HAS_TKINTER = False

from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles

def get_resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def get_external_path(relative_path):
    """ Get path to resource outside the executable bundle (portable) """
    if getattr(sys, 'frozen', False):
        # The application is frozen (running as an .exe)
        base_path = os.path.dirname(sys.executable)
    else:
        # The application is running in a normal Python environment
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

app = FastAPI()

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    print(f"Validation Error: {exc.errors()}")
    return JSONResponse(status_code=422, content={"detail": exc.errors()})

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration for persistence and resources
# We use get_external_path for things that should be editable/visible by the user (templates, output)
# We use get_resource_path for things bundled inside (frontend dist)
BASE_DIR = os.path.abspath(".")
UPLOADS_DIR = get_external_path("uploads")
OUTPUT_DIR = get_external_path("output")
TEMPLATES_DIR = get_external_path("templates")

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

class ColorRow(BaseModel):
    material: str
    pantone: str
    color: str
    pasada: Optional[str] = ""
    maquina: Optional[str] = ""
    temp: Optional[str] = ""
    t_expo: Optional[str] = ""

class FichaData(BaseModel):
    # AZUL
    marca: str
    dis_grafico: str
    dis_moda: str
    campana: str
    anio: str
    linea: str
    proceso: str
    tipo_logo: str
    pieza: str
    foil: str
    # ROJO
    referencia: str
    tela_base: str
    num_piezas: str
    porc_cubrimiento: str
    alta_densi: str
    fecha_ficha: str
    ancho_cm: str
    alto_cm: str
    num_estampados: str
    observaciones: str
    output_folder: Optional[str] = ""
    # IMAGES (base64)
    img_ropero_cad: str
    img_pantallazo_molde: str
    # COLORS
    colores: List[ColorRow]

# CONTINUO MODELS
class ContinuoGlobal(BaseModel):
    referencia: str
    marca: str
    campana: str
    linea: str
    dis_moda: str
    dis_grafico: str
    proceso: str
    observaciones: str
    imagen_pattern_base64: str
    output_folder: Optional[str] = ""

class TelaContinuo(BaseModel):
    nombre_tela: str
    composicion: str
    pantones: Optional[List[str]] = []

class ContinuoRequest(BaseModel):
    datos_globales: ContinuoGlobal
    telas_seleccionadas: List[TelaContinuo]

TEMPLATE_LOCALIZADO = os.path.join(TEMPLATES_DIR, "Ficha_Desrollo_Estampacion_V01.xlsm")
TEMPLATE_CONTINUA = os.path.join(TEMPLATES_DIR, "Ficha_Continua_V01.xlsm")

CACHED_OPCIONES = None

@app.get("/api/select-folder")
def select_folder():
    """ Opens a native folder selection dialog """
    if not HAS_TKINTER:
        return {"error": "Folder selection is not available on this server environment."}
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        root.attributes('-topmost', True) # Bring to front
        folder_selected = filedialog.askdirectory()
        root.destroy()
        return {"path": folder_selected}
    except Exception as e:
        return {"error": f"Folder selection not available: {str(e)}"}

@app.get("/api/config")
def get_config():
    is_render = os.environ.get("RENDER") is not None
    return {
        "is_local": not is_render and HAS_TKINTER
    }

@app.get("/api/opciones")
def get_opciones():
    global CACHED_OPCIONES
    try:
        if not os.path.exists(TEMPLATE_LOCALIZADO):
            raise HTTPException(status_code=500, detail=f"Template not found at {TEMPLATE_LOCALIZADO}")
            
        wb = openpyxl.load_workbook(TEMPLATE_LOCALIZADO, read_only=True, data_only=True)
        if "DATOS" not in wb.sheetnames:
            raise HTTPException(status_code=500, detail="Sheet 'DATOS' not found in template")
        
        ws = wb["DATOS"]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break
        if not headers: return {}

        def get_col_index(name):
            try: return headers.index(name)
            except ValueError: return -1

        mapping = {
            "marca": get_col_index("MARCA"),
            "linea": get_col_index("LINEA"),
            "dis_graf": get_col_index("DIS. GRAF"),
            "dis_mod": get_col_index("DIS.MOD"),
            "campana": get_col_index("CAMPAÑA"),
            "anio": get_col_index("AÑO"),
            "proceso": get_col_index("PROCESO"),
            "tipo_logo": get_col_index("TIPO LOGO"),
            "pieza": get_col_index("PIEZA"),
            "foil": get_col_index("FOIL"),
            "material": get_col_index("MATERIAL"),
            "tela_base": 1,        # Column B
            "num_estampado": 8     # Column I
        }

        opciones = {k: set() for k in mapping.keys()}
        empty_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                empty_count += 1
                if empty_count > 10: break
                continue
            empty_count = 0
            for key, idx in mapping.items():
                if idx != -1 and idx < len(row) and row[idx] is not None:
                    opciones[key].add(str(row[idx]))

        result = {k: sorted(list(v)) for k, v in opciones.items()}
        CACHED_OPCIONES = result
        return result
    except Exception as e:
        print(f"Error reading options: {e}")
        raise HTTPException(status_code=500, detail=str(e))

def get_composition_mapping():
    try:
        if not os.path.exists(TEMPLATE_CONTINUA):
            return {}
        wb = openpyxl.load_workbook(TEMPLATE_CONTINUA, read_only=True, data_only=True)
        # Try to find a sheet with composition mapping or the second sheet
        sheet_name = '  '
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
            
        ws = wb[sheet_name]
        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3 and row[1] and row[2]:
                mapping[str(row[1]).strip()] = str(row[2]).strip()
        return mapping
    except Exception as e:
        print(f"Error reading composition mapping: {e}")
        return {}

def save_base64_image(base64_str, folder, filename):
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
    img_data = base64.b64decode(base64_str)
    filepath = os.path.join(folder, filename)
    with open(filepath, "wb") as f:
        f.write(img_data)
    return filepath

def generar_ficha_core(data: FichaData) -> tuple[str, str]:
    ficha_id = str(uuid.uuid4())
    ficha_folder = os.path.join(UPLOADS_DIR, ficha_id)
    os.makedirs(ficha_folder, exist_ok=True)

    path_ropero = save_base64_image(data.img_ropero_cad, ficha_folder, "ropero.png")
    path_molde = save_base64_image(data.img_pantallazo_molde, ficha_folder, "molde.png")

    wb = openpyxl.load_workbook(TEMPLATE_LOCALIZADO, keep_vba=True)
    ws = wb["FICHA"]

    MAPEO_FICHA = {
        "marca": "B5", "dis_grafico": "D5", "dis_moda": "H5", "campana": "E3",
        "anio": "G3", "linea": "I3", "proceso": "I4", "tipo_logo": "K5",
        "pieza": "B7", "foil": "K7",
        "referencia": "C3", "tela_base": "B6", "num_piezas": "F7",
        "porc_cubrimiento": "I7", "alta_densi": "B8", "fecha_ficha": "K8",
        "ancho_cm": "B21", "alto_cm": "B22", "num_estampados": "B23", 
        "observaciones": "H31"
    }

    from openpyxl.cell.cell import MergedCell
    def write_to_cell(sheet, coord, value):
        cell = sheet[coord]
        if isinstance(cell, MergedCell):
            for m_range in sheet.merged_cells.ranges:
                if coord in m_range:
                    sheet.cell(row=m_range.min_row, column=m_range.min_col).value = value
                    return
        cell.value = value

    for key, cell in MAPEO_FICHA.items():
        if hasattr(data, key):
            val = getattr(data, key)
            write_to_cell(ws, cell, val)

    def add_image(ws, path, cell, size=(300, 300)):
        img = Image(path)
        with PILImage.open(path) as pil_img:
            width, height = pil_img.size
            ratio = min(size[0]/width, size[1]/height)
            img.width = width * ratio
            img.height = height * ratio
        img.anchor = cell
        ws.add_image(img)

    add_image(ws, path_ropero, "B12", size=(280, 350))
    add_image(ws, path_molde, "H12", size=(550, 450))

    start_row = 30
    for i, color in enumerate(data.colores[:12]):
        row_idx = start_row + i
        write_to_cell(ws, f"A{row_idx}", color.material)
        write_to_cell(ws, f"B{row_idx}", color.pantone)
        write_to_cell(ws, f"C{row_idx}", color.color)

    clean_ref = "".join(x for x in data.referencia if x.isalnum())
    filename = f"{clean_ref}.xlsm" if clean_ref else "Ficha.xlsm"
    out_dir = data.output_folder if data.output_folder and os.path.isdir(data.output_folder) else OUTPUT_DIR
    out_path = os.path.join(out_dir, filename)
    
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)

    if os.path.exists(path_ropero): os.remove(path_ropero)
    if os.path.exists(path_molde): os.remove(path_molde)

    return out_path, filename


def generar_continuo_core(req: ContinuoRequest) -> tuple[str, int]:
    data = req.datos_globales
    telas = req.telas_seleccionadas
    MAPEO_CELDAS = {
        "marca": "F1", "campana": "F2", "linea": "F3",
        "dis_moda": "C4", "dis_grafico": "F4",
        "proceso": "C5", "referencia": "D7",
        "observaciones": "A20"
    }
    
    # Determine output directory
    out_root = data.output_folder if data.output_folder and os.path.isdir(data.output_folder) else OUTPUT_DIR
    folder_name = f"Continuo_{data.referencia}"
    folder_path = os.path.join(out_root, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    
    # Clean base64 header if present
    base64_str = data.imagen_pattern_base64
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
        
    path_pattern = save_base64_image(base64_str, folder_path, "pattern_base.png")
    
    comp_mapping = get_composition_mapping()

    for tela in telas:
        wb = openpyxl.load_workbook(TEMPLATE_CONTINUA, keep_vba=True)
        ws = wb["FICHA"] if "FICHA" in wb.sheetnames else wb.active
        
        from openpyxl.cell.cell import MergedCell
        def write_cell(sheet, coord, val):
            cell = sheet[coord]
            if isinstance(cell, MergedCell):
                for m_range in sheet.merged_cells.ranges:
                    if coord in m_range:
                        sheet.cell(row=m_range.min_row, column=m_range.min_col).value = val
                        return
            cell.value = val

        for key, cell in MAPEO_CELDAS.items():
            val = getattr(data, key)
            write_cell(ws, cell, val)
        
        write_cell(ws, "H3", tela.nombre_tela)
        
        try:
            img = Image(path_pattern)
            with PILImage.open(path_pattern) as pil_img:
                w, h = pil_img.size
                target_w, target_h = 380, 280 
                ratio = min(target_w/w, target_h/h)
                img.width = w * ratio
                img.height = h * ratio
            ws.add_image(img, "B11")
        except Exception as img_e:
            print(f"Error insertando imagen: {img_e}")
        
        if data.proceso == "SUBLIMACION":
            for row in range(11, 21): write_cell(ws, f"H{row}", None)
        else: # ESTAMPACION
            if tela.pantones:
                for i, p in enumerate(tela.pantones):
                    if i < 10: write_cell(ws, f"H{11+i}", p)
        
        safe_tela = "".join(x for x in tela.nombre_tela if x.isalnum())[:30]
        safe_ref = "".join(x for x in data.referencia if x.isalnum())[:20]
        filename = f"F_{safe_ref}_{safe_tela}.xlsm"
        
        save_path = os.path.join(folder_path, filename)
        wb.save(save_path)
        wb.close()
        
    return folder_path, len(telas)


# BACKGROUND TASK QUEUE AND WORKER
TASKS_DB = {}
task_queue = queue.Queue()

def task_worker():
    while True:
        task_id = task_queue.get()
        if task_id is None:
            break
        
        task = TASKS_DB.get(task_id)
        if not task:
            task_queue.task_done()
            continue
        
        task["estado"] = "processing"
        task["inicio"] = datetime.now().isoformat()
        
        try:
            if task["tipo"] == "LOCALIZADO":
                out_path, filename = generar_ficha_core(task["data"])
                task["estado"] = "completed"
                task["out_path"] = out_path
                task["filename"] = filename
            elif task["tipo"] == "CONTINUO":
                folder_path, count = generar_continuo_core(task["data"])
                import shutil
                # Create zip archive of the folder
                zip_path = shutil.make_archive(folder_path, 'zip', folder_path)
                task["estado"] = "completed"
                task["out_path"] = folder_path
                task["zip_path"] = zip_path
                task["zip_filename"] = f"{os.path.basename(folder_path)}.zip"
                task["message"] = f"Generadas {count} fichas."
        except Exception as e:
            traceback.print_exc()
            task["estado"] = "failed"
            task["error"] = str(e)
        finally:
            task["fin"] = datetime.now().isoformat()
            task_queue.task_done()

# Start background thread
worker_thread = threading.Thread(target=task_worker, daemon=True)
worker_thread.start()


# ORIGINAL SYNCHRONOUS ENDPOINTS (UNMODIFIED BEHAVIOR)
@app.post("/api/generar-ficha")
def generar_ficha(data: FichaData):
    try:
        out_path, filename = generar_ficha_core(data)
        return FileResponse(path=out_path, filename=filename, media_type='application/vnd.ms-excel.sheet.macroEnabled.12')
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generar-continuo")
def generar_continuo(req: ContinuoRequest):
    try:
        folder_path, count = generar_continuo_core(req)
        return {"success": True, "message": f"Generadas {count} fichas.", "folder_path": folder_path}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# NEW ASYNCHRONOUS QUEUE ENDPOINTS
@app.post("/api/generar-ficha-async")
def generar_ficha_async(data: FichaData):
    try:
        task_id = str(uuid.uuid4())
        TASKS_DB[task_id] = {
            "id": task_id,
            "tipo": "LOCALIZADO",
            "referencia": data.referencia,
            "estado": "queued",
            "timestamp": datetime.now().isoformat(),
            "data": data
        }
        task_queue.put(task_id)
        return {"task_id": task_id, "status": "queued"}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generar-continuo-async")
def generar_continuo_async(req: ContinuoRequest):
    try:
        task_id = str(uuid.uuid4())
        TASKS_DB[task_id] = {
            "id": task_id,
            "tipo": "CONTINUO",
            "referencia": req.datos_globales.referencia,
            "estado": "queued",
            "timestamp": datetime.now().isoformat(),
            "data": req
        }
        task_queue.put(task_id)
        return {"task_id": task_id, "status": "queued"}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tasks")
def get_tasks():
    result = []
    # Sort tasks by creation time descending
    for t_id, t in sorted(TASKS_DB.items(), key=lambda x: x[1]["timestamp"], reverse=True):
        result.append({
            "id": t["id"],
            "tipo": t["tipo"],
            "referencia": t["referencia"],
            "estado": t["estado"],
            "timestamp": t["timestamp"],
            "inicio": t.get("inicio"),
            "fin": t.get("fin"),
            "filename": t.get("filename"),
            "message": t.get("message"),
            "error": t.get("error")
        })
    return result


@app.get("/api/tasks/{task_id}/download")
def download_task_file(task_id: str):
    task = TASKS_DB.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if task["estado"] != "completed":
        raise HTTPException(status_code=400, detail="Task is not completed yet")
    
    if task["tipo"] == "LOCALIZADO":
        out_path = task.get("out_path")
        filename = task.get("filename", "Ficha.xlsm")
        media_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
    elif task["tipo"] == "CONTINUO":
        out_path = task.get("zip_path")
        filename = task.get("zip_filename", "Continuo.zip")
        media_type = 'application/zip'
    else:
        raise HTTPException(status_code=400, detail="Invalid task type")
        
    if not out_path or not os.path.exists(out_path):
        raise HTTPException(status_code=404, detail="File not found on server")
        
    return FileResponse(path=out_path, filename=filename, media_type=media_type)

# Serving React Frontend
DIST_PATH = os.path.join(os.path.abspath("."), "frontend_dist")
if not os.path.exists(DIST_PATH):
    DIST_PATH = get_resource_path("dist")
if not os.path.exists(DIST_PATH):
    DIST_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend_dist")

if os.path.exists(DIST_PATH):
    app.mount("/", StaticFiles(directory=DIST_PATH, html=True), name="static")
else:
    print(f"Warning: Static frontend directory not found. Checked: {DIST_PATH}")

if __name__ == "__main__":
    import uvicorn
    # Use RENDER port or default to 8000
    port = int(os.environ.get("PORT", 8000))
    # Bind to 0.0.0.0 to allow external routing in cloud environments
    uvicorn.run(app, host="0.0.0.0", port=port)
